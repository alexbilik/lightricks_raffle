import traceback, sys, random, argparse, os
from gooey import Gooey
import logging
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from collections import namedtuple

class FilterAboveOrEqualToLevel(object):
    def __init__(self, level):
        self.__level = level

    def filter(self, record):
        return record.levelno >= self.__level


class FilterBelowOrEqualToLevel(object):
    def __init__(self, level):
        self.__level = level

    def filter(self, record):
        return record.levelno <= self.__level

# Globals:
PEOPLE_CHOICE_TUPLE = namedtuple('PEOPLE_CHOICE_TUPLE', ['NAME', 'ROW', 'CHOICES', 'DECISION'])

DEFAULT_DEBUG_FILE_NAME = 'lightricks_raffle_debug.log'
logger = None


def add_to_exception_string(message):
    global exception_str

    exception_str += 'Exception message: {}\n\n{}'.format(message, traceback.format_exc())


def setup_logging(name, debug_file_location, stdout_log_level):
    """
    Description: Set logging name, level and config.
    :param name: Name of the logger
    :param debug_file_location: Location of debug log
    :return:
    """
    global logger
    fmt = '%(asctime)s %(name)12s (%(lineno)4d): %(levelname)-8s %(message)s'
    logger = logging.getLogger(name)
    logger.setLevel(logging.DEBUG)
    logger.propagate = False
    log_fmt = logging.Formatter(fmt)

    if debug_file_location is not None:
        output_to_file_handler = logging.FileHandler(filename=debug_file_location if debug_file_location else
        os.path.join(os.getcwd(), DEFAULT_DEBUG_FILE_NAME))
        output_to_file_handler.setFormatter(log_fmt)
        output_to_file_handler.setLevel(logging.DEBUG)
        logger.addHandler(output_to_file_handler)

    output_to_screen_handler = logging.StreamHandler(stream=sys.stdout)
    output_to_screen_stderr_handler = logging.StreamHandler(stream=sys.stderr)
    output_to_screen_handler.setFormatter(log_fmt)
    output_to_screen_stderr_handler.setFormatter(log_fmt)
    output_to_screen_handler.setLevel(stdout_log_level)
    output_to_screen_stderr_handler.setLevel(logging.ERROR)
    output_to_screen_handler.addFilter(FilterBelowOrEqualToLevel(logging.WARNING))
    output_to_screen_stderr_handler.addFilter(FilterAboveOrEqualToLevel(logging.ERROR))
    logger.addHandler(output_to_screen_handler)
    logger.addHandler(output_to_screen_stderr_handler)

    return logger



def get_stdout_log_level(given_level):
    if given_level == "debug":
        return logging.DEBUG
    elif given_level == "warning":
        return logging.WARNING

    return logging.INFO


def get_workers_decisions_and_inventory_dict(projects_xlsx):
    inventory = {}
    people_choices = []

    workbook = load_workbook(filename=projects_xlsx, data_only=True)
    logger.debug(workbook.sheetnames)
    if not u'Form Responses 1' in workbook.sheetnames:
        logger.error("u'Form Responses 1' not in sheetnames")
    responses_sheet = workbook[u'Form Responses 1']
    if not u'Inventory' in workbook.sheetnames:
        logger.error("u'Inventory' not in sheetnames")
    inventory_sheet = workbook[u'Inventory']

    for item in inventory_sheet['B'][1:]:
        prize_name = item.internal_value
        prize_count = inventory_sheet['C'][item.row - 1].internal_value
        logger.info('Found prize: {} count: {}'.format(prize_name, prize_count))
        if prize_name is None or prize_count is None:
            break
        inventory[prize_name] = prize_count


    for person in responses_sheet['B'][1:]:
        name = person.internal_value
        choice1 = responses_sheet['C'][person.row - 1].internal_value
        choice2 = responses_sheet['D'][person.row - 1].internal_value
        choice3 = responses_sheet['E'][person.row - 1].internal_value
        logger.info('Person: {} Choice1: {} Choice2: {} Choice3: {}'.format(name, choice1, choice2, choice3))
        if name is None:
            break
        people_choices.append(PEOPLE_CHOICE_TUPLE(NAME=name,
                                                  ROW=person.row,
                                                  CHOICES=[choice1, choice2, choice3], DECISION={'won':False, 'prize':None, 'selection':None}))

    return people_choices, inventory, workbook


def get_all_selected_inv_item_people(people_choices, choice_num, inv_item):
    selected_people = [s for s in people_choices if s.CHOICES[choice_num] == inv_item]

    return selected_people



def make_raffle(people_choices, inventory):
    new_people_choices = []
    for choice_num in range(3):
        logger.info('Checking choice: {}'.format(choice_num))

        for inv_item in inventory:
            selected_people = get_all_selected_inv_item_people(people_choices, choice_num, inv_item)
            if len(selected_people) == 0:
                logger.info('No people selected item: {} in choice: {}'.format(inv_item, choice_num))
                continue

            inv_item_total = inventory[inv_item]
            if inv_item_total > len(selected_people):
                num_to_choose =  len(selected_people)
            else:
                num_to_choose = inv_item_total
            logger.info('inv_item_total: {} num_to_choose: {} len(selected_people): {}'.format(inv_item_total, num_to_choose, len(selected_people)))
            if inv_item_total <= 0:
                logger.info('No more items of type: {}'.format(inv_item))
                continue
            selected_people_nums = set(random.sample(range(len(selected_people)), num_to_choose))
            logger.info('item: {}, selected_people_nums: {}'.format(inv_item, selected_people_nums))

            for i in selected_people_nums:
                new_people_choices.append(selected_people[i])
                new_people_choices[-1].DECISION['won'] = True
                new_people_choices[-1].DECISION['prize'] = inv_item
                new_people_choices[-1].DECISION['selection'] = choice_num + 1
                inventory[inv_item] -= 1
                logger.info('person: {} won: {} choice: {}'.format(new_people_choices[-1].NAME, new_people_choices[-1].DECISION['prize'], new_people_choices[-1].DECISION['selection']))
                logger.info('quantity for item: {} is: {}'.format(inv_item, inventory[inv_item]))
                people_choices.remove(selected_people[i])

    return new_people_choices, inventory


def write_results_to_excel(new_people_choices, inventory, workbook, raffle_excel_output_file_path):
    responses_sheet = workbook[u'Form Responses 1']
    inventory_sheet = workbook[u'Inventory']

    responses_sheet.cell(row=1, column=column_index_from_string('J')).value = 'Won'
    responses_sheet.cell(row=1, column=column_index_from_string('K')).value = 'Prize'
    responses_sheet.cell(row=1, column=column_index_from_string('L')).value = 'Choice'
    for people_choice in new_people_choices:
        responses_sheet.cell(row=people_choice.ROW, column=column_index_from_string('J')).value = people_choice.DECISION['won']
        responses_sheet.cell(row=people_choice.ROW, column=column_index_from_string('K')).value = people_choice.DECISION['prize']
        responses_sheet.cell(row=people_choice.ROW, column=column_index_from_string('L')).value = people_choice.DECISION['selection']

    inventory_sheet.cell(row=1, column=column_index_from_string('E')).value = 'Inventory leftovers'
    for i, inv_item in enumerate(inventory):
        inventory_sheet.cell(row=i+2, column=column_index_from_string('E')).value = inventory[inv_item]


    workbook.save(raffle_excel_output_file_path)


@Gooey
def main():
    global logger, exit_status
    help_level_desc = "Level of the log printed to screen. Error and Critical level logs are unsupported, and " \
                      "will be printed anyway."
    parser = argparse.ArgumentParser(description="")
    parser.add_argument('-d', '--debug', help="Location of debug file log. Elaborated messages for the user.", type=str, default=DEFAULT_DEBUG_FILE_NAME)
    parser.add_argument('-x', '--raffle_excel_file_path', help="Raffle results Excel full file path", required=True, type=str)
    parser.add_argument('-o', '--raffle_excel_output_file_path', help="Raffle results Excel output full file path", required=True, type=str)
    parser.add_argument('-l', "--log_level", help=help_level_desc, choices=["debug", "info", "warning"], default="debug")
    args = parser.parse_args()
    logger = setup_logging("TFS queries Connection", args.debug, get_stdout_log_level(args.log_level))

    try:
        logger.info('Input excel: {}'.format(args.raffle_excel_file_path.replace('"', '').replace('\'', '')))
        logger.info('Output excel: {}'.format(args.raffle_excel_output_file_path.replace('"', '').replace('\'', '')))
        people_choices, inventory, workbook = get_workers_decisions_and_inventory_dict(projects_xlsx=args.raffle_excel_file_path.replace('"', '').replace('\'', ''))

        new_people_choices, inventory = make_raffle(people_choices, inventory)
        write_results_to_excel(new_people_choices, inventory, workbook, args.raffle_excel_output_file_path.replace('"', '').replace('\'', ''))
    except Exception as e:
        logger.error(traceback.format_exc())
        exit(1)

    logger.info('All Done!')

if __name__ == "__main__":
    main()
